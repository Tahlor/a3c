import numpy as np
import pickle
import os
import tensorflow as tf

import fnmatch

#tf.random_uniform_initializer(-1e-3, 1e-3)
# variance_scaling_initializer 
# random_normal_initializer
#archipack_default_initializer = tf.variance_scaling_initializer(1)
archipack_default_initializer = tf.random_normal_initializer(stddev=0.02)

def one_hot(tensor, num_of_classes = None):
    t = (np.array(tensor, dtype=np.int32))
    if t.shape >1:
        t = t.reshape(-1)
    n_values = np.max(t) + 1
    one_hot_targets = np.eye(n_values, num_of_classes)[t]
    return one_hot_targets

def linear2():
    f1 = fc(z, out_size=4*4*256, name="fc")
    h1 = tf.reshape(f1, [-1, 4, 4, 256], name="h1")

def linear(input_, output_size, scope=None, stddev=0.02, bias_start=0.0):
    #shape = input_.get_shape().as_list()
    shape = np.shape(input_)
    with tf.variable_scope(scope or "Linear"):
        matrix = tf.get_variable("Matrix", [shape[1], output_size], tf.float32,
            tf.random_normal_initializer(stddev=stddev))
        bias = tf.get_variable("bias", [output_size],
            initializer=tf.constant_initializer(bias_start))
        return tf.matmul(input_, matrix) + bias
    
def find_files(base, pattern):
    '''Return list of files matching pattern in base folder.'''
    #matching_files = [n for n in fnmatch.filter(os.listdir(base), pattern) if os.path.isfile(os.path.join(base, n))]
    matching_files_and_folders = fnmatch.filter(os.listdir(base), pattern)
    return len(matching_files_and_folders)>0

def createLogDir(name = "", force_numerical_ordering = True):
    n = 1

    # Add padding
    if name != "" and name[0] != " ":
        name = " " + name

    # Check for existence
    basepath = "./tf_logs"
    if not os.path.exists(basepath):
        os.mkdir(basepath)

    if force_numerical_ordering:
        while find_files(basepath, str(n) + " *") or os.path.exists(os.path.join(basepath, str(n)    )) :
            n += 1
    else:
        while os.path.exists(os.path.join(basepath, str(n) + name )):
            n += 1

    # Create
    logdir = os.path.join(basepath, str(n) + name)
    os.mkdir(logdir)
    training_accuracy_list = []
    print(logdir)
    return logdir

def shuffleDataAndLabelsInPlace ( arr1, arr2, seed = None):
    from numpy.random import RandomState
    import sys
    if seed is None:
        seed = np.random.randint(0, sys.maxsize/10**10)
    prng = RandomState(seed)
    prng.shuffle(arr1)
    prng = RandomState(seed)
    prng.shuffle(arr2)
        
def unpickleCIFAR( file ):
    fo = open(file, 'rb')
    dict = pickle.load(fo, encoding = 'latin-1')
    fo.close()
    return dict

def importCIFAR(filename = 'cifar-10-batches-py/data_batch_1'): 
    data = unpickle( filename )
    features = data['data']
    labels = data['labels']
    labels = np.atleast_2d( labels ).T
    return labels, features

def importCIFARall():
    filename = 'cifar-10-batches-py/data_batch_1'
    data = unpickle(filename)
    features = data['data']
    labels = data['labels']
    for i in range(2,5):
        filename = 'cifar-10-batches-py/data_batch_' + str(i)
        data = unpickle(filename)
        labels = np.append(labels, data['labels'])
        features = np.vstack([features, data['data']])

    labels = np.atleast_2d( labels ).T
    print(features.shape)
    print(labels.shape)
    return labels, features

def splitTrainingTest(percent_training, features, labels):
    assert features.shape[0] == labels.shape[0]
    count = features.shape[0]
    percent_training = percent_training/100 if percent_training > 1 else percent_training
    training_size = int(count*percent_training)
    training_features  = features[:training_size]
    test_features      = features[training_size:]
    training_labels    = labels[:training_size]
    test_labels        = labels[training_size:]
    return training_features, training_labels, test_features, test_labels

def convNew( x, output_shape = [], filter_size=3, stride=2, num_filters=64, is_output=False, name="conv", activation = "RELU", initializer = None, batch_norm = False, reuse = False):
    if is_output == True:
        activation_fn = None
    elif activation == "RELU":
        activation_fn = tf.nn.relu     
    else:
        activation_fn = None
    
    normalizer_fn = tf.contrib.layers.batch_norm if batch_norm else None
    
    with tf.variable_scope(name, reuse = reuse):
        conv = tf.contrib.layers.convolution2d(inputs=x,
            num_outputs=num_filters, # 4
            kernel_size=[filter_size, filter_size], # [1,3]
            stride=[stride,stride],
            padding='SAME',
            rate=1,
            activation_fn=activation_fn,
            normalizer_fn=normalizer_fn,
            normalizer_params=None,
            weights_initializer=tf.contrib.layers.xavier_initializer(dtype=tf.float32),
            biases_initializer=tf.zeros_initializer,
            trainable=True,
            scope='cnn',
            reuse = reuse
            ) 
        if activation == "leakyRELU":
            conv = leakyRELU(conv, .2)
    return conv, 0

 
def conv( x, output_shape = [], filter_size=3, stride=2, num_filters=64, is_output=False, name="conv", activation = "RELU", initializer = None, batch_norm = False):
    x_channels = np.shape(x)[3]

    # Make sure user knows what their doing
    if output_shape != []:
        #print(output_shape)
        #print(np.shape(x))
        #assert output_shape[0] == np.shape(x)[0] #batch_size, often ? at this point
        assert int(output_shape[1]) == int(np.shape(x)[1])/stride
        assert int(output_shape[2]) == int(np.shape(x)[2])/stride
        assert int(output_shape[3]) == num_filters
        
    with tf.name_scope(name) as scope:
        if initializer == None:
            archipack_default_initializer
        W = tf.get_variable(name+"_W", [filter_size, filter_size, x_channels, num_filters], np.float32, initializer=archipack_default_initializer )        
        b = tf.get_variable(name+"_b", [num_filters], np.float32, initializer=archipack_default_initializer)

        #W = tf.Variable( 1e-3*np.random.randn( filter_size, filter_size, x_channels, num_filters ).astype(np.float32), name="W" )
        #b = tf.Variable( 1e-3*np.random.randn( num_filters).astype(np.float32), name="b" )
        op = tf.nn.conv2d(x, W, strides=[1, stride, stride, 1], padding='SAME')
        op = tf.nn.bias_add(op, b)

    if batch_norm:
        op = tf.layers.batch_normalization(op)

        
    if not is_output:
        if activation == "leakyRELU":
            op = leakyRELU(op)
        else:
            op = tf.nn.relu(op)
    return op, W

def leakyRELU(x, alpha=0.2):
    return tf.maximum(alpha*x, x)

#shape of input = [batch, in_height, in_width, in_channels]
#shape of filter = [filter_height, filter_width, in_channels, out_channels]

def fc( x, out_size=50, name="fc" , is_output=False, batch_size = 1, batch_norm = False):    
    x_dim = x.get_shape().as_list()[1]
    
    with tf.name_scope(name) as scope:
        W = tf.get_variable(name+"_W", [out_size, x_dim], np.float32, initializer=archipack_default_initializer )        
        b = tf.get_variable(name+"_b", [out_size, 1], np.float32, initializer=archipack_default_initializer)

        #W = tf.Variable( 1e-3*np.random.randn(out_size, x_dim).astype(np.float32), name="W")
        #b = tf.Variable( 1e-3*np.random.randn(out_size, 1).astype(np.float32), name="b" )        
        
        # should automagically take care of batching 
        # op = tf.matmul(W, x) + b
        op = tf.einsum('ij,bjk->bik', W, x) + b
        
        #op = tf.nn.bias_add(op, b)

    if batch_norm:
        op = tf.layers.batch_normalization(op)

        
    if not is_output:
        op = tf.nn.relu(op)
        
    return op

def max_pool_2x2(x):
  """max_pool_2x2 downsamples a feature map by 2X."""
  return tf.nn.max_pool(x, ksize=[1, 2, 2, 1], strides=[1, 2, 2, 1], padding='SAME')


def deconv( x, output_shape = [1,32,32,3], filter_size=3, stride=2, num_filters=64, is_output=False, name="deconv" , filter_stride = 1, batch_norm = False):
   # tf.layers.conv2d_transpose is better for batching
    x_channels = np.shape(x)[3]
    with tf.name_scope(name) as scope:
        W = tf.get_variable(name+"_W", [filter_size, filter_size, num_filters, x_channels], np.float32, initializer=archipack_default_initializer )        
        b = tf.get_variable(name+"_b", [num_filters], np.float32, initializer=archipack_default_initializer )
        
        #W = tf.Variable( 1e-3*np.random.randn( filter_size, filter_size, num_filters, x_channels).astype(np.float32), name="W")
        #b = tf.Variable( 1e-3*np.random.randn( num_filters).astype(np.float32), name="b")
        #print(W.get_shape().as_list())
        #print(x.get_shape().as_list())
        op = tf.nn.conv2d_transpose(x, W, output_shape, strides=[1, stride, stride, filter_stride], padding='SAME', data_format='NHWC', name="deconv")
        op = tf.nn.bias_add(op, b)
    
    if batch_norm:
        op = tf.layers.batch_normalization(op)
        
    if not is_output:
        op = tf.nn.relu(op)
    return op, W



### UNFINISHED
def max_pool( x, out_size=50, name="max_pool" , is_output=False):


    x_dim = np.shape(x)[0]
    x_dim = x.get_shape().as_list()[0]

    with tf.name_scope(name) as scope:
        W = tf.Variable( 1e-3*np.random.randn(out_size, x_dim).astype(np.float32), name="W" )
        b = tf.Variable( 1e-3*np.random.randn(out_size, 1).astype(np.float32), name="b" )
        op = tf.matmul(W, x) + b
        #op = tf.nn.bias_add(op, b)
        
    if not is_output:
        op = tf.nn.relu(op)
        
    return op


def conv2( x, filter_size=3, stride=2, num_filters=64, is_output=False, name="conv" ):

    '''
    x is an input tensor
    Declare a name scope using the "name" parameter
    Within that scope:
      Create a W filter variable with the proper size
      Create a B bias    with the proper size
      Convolve x with W by calling the tf.nn.conv2d function
      Add the bias
      If is_output is False,
        Call the tf.nn.relu function
      Return the final op
    '''
    x_channels = np.shape(x)[3]
    with tf.name_scope(name) as scope:
        W = tf.Variable( 1e-3*np.random.randn( filter_size, filter_size, x_channels, num_filters ).astype(np.float32), name="W" )
        #b = tf.Variable( 1e-3*np.random.randn( 1, 1, 1, num_filters ).astype(np.float32), name="b" )
        b = tf.Variable( 1e-3*np.random.randn( num_filters).astype(np.float32), name="b" )
        op = tf.nn.conv2d(x, W, strides=[1, stride, stride, 1], padding='SAME')
        op = tf.nn.bias_add(op, b)
       
    if not is_output:
        op = tf.nn.relu(op)
    return op
#shape of input = [batch, in_height, in_width, in_channels]
#shape of filter = [filter_height, filter_width, in_channels, out_channels]

def fc2( x, out_size=50, name="fc" , is_output=False):

    '''
    x is an input tensor - we expect a vector
    Declare a name scope using the "name" parameter
    Within that scope:
      Create a W filter variable with the proper size
      Create a B bias variable with the proper size
      Multiply x by W and add b
      If is_output is False,
        Call the tf.nn.relu function
      Return the final op
    '''
    x_dim = np.shape(x)[0]
    x_dim = x.get_shape().as_list()[0]

    with tf.name_scope(name) as scope:
        W = tf.Variable( 1e-3*np.random.randn(out_size, x_dim).astype(np.float32), name="W" )
        b = tf.Variable( 1e-3*np.random.randn(out_size, 1).astype(np.float32), name="b" )
        op = tf.matmul(W, x) + b
        #op = tf.nn.bias_add(op, b)
        
    if not is_output:
        op = tf.nn.relu(op)
        
    return op

def writeCSV(theList, filepath, append = False):
    import csv
    writemode = "wb" if not append else "a"
    with open(filepath, writemode) as f:
            writer = csv.writer(f)
            writer.writerows(theList)

def writeXLSX(theList, filepath, sheet = None):
    import xlsxwriter
    workbook = xlsxwriter.Workbook(filepath)
    worksheet = workbook.add_worksheet(sheet)
    for i, row in enumerate(theList):
        for j,column in enumerate(row):    
            worksheet.write(i, j, column)
    workbook.close()

def modifyXLSX(theList, filepath, sheet = None):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(filepath)
    except:
        wb = openpyxl.Workbook()

    #ws = wb.get_sheet_by_name(sheet)
    ws = wb.create_sheet(sheet)

    for i, row in enumerate(theList):
        for j,value in enumerate(row):    
            ws.cell(row = i+1, column = j+1).value = value
    wb.save(filepath)

# NOT FINISHED
def modifyXLSX2(theList, filepath, sheet= None):
    import pandas as pd
    wb = Workbook(Existing_file)
    df = Range(Anchor).table.value
    df = pd.DataFrame(df) # into Pandas DataFrame
    df['sum'] = df.sum(axis= 1)
    Range(Anchor).value = df.values

    
if __name__ == '__main__':
    #importCIFARall()
    logdir = createLogDir()
    pass
    
"""normalizer_fn = tf.contrib.layers.batch_norm

D = 5
kernel_height = 1
kernel_width = 3
F = 4
x = tf.placeholder(tf.float32, shape=[None,1,D,1], name='x-input') #[M, 1, D, 1]
conv = tf.contrib.layers.convolution2d(inputs=x,
    num_outputs=F, # 4
    kernel_size=[kernel_height, kernel_width], # [1,3]
    stride=[1,1],
    padding='VALID',
    rate=1,
    activation_fn=tf.nn.relu,
    normalizer_fn=normalizer_fn,
    normalizer_params=None,
    weights_initializer=tf.contrib.layers.xavier_initializer(dtype=tf.float32),
    biases_initializer=tf.zeros_initializer,
    trainable=True,
    scope='cnn'
)
"""
